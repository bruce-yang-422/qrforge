import io
import os
import re
import csv
import logging
import shutil
import importlib.util
from datetime import datetime
import segno
from PIL import Image
from openpyxl import load_workbook


# =========================
# CONFIG 設定載入
# =========================
# Loads configuration from config/settings.py at runtime using importlib,
# so the file can contain Python comments without JSON parsing restrictions.
# 使用 importlib 動態載入設定檔，讓設定檔可使用 Python 註解。

def load_config():
    spec = importlib.util.spec_from_file_location("settings", "config/settings.py")
    if spec is None or spec.loader is None:
        raise RuntimeError("無法載入 config/settings.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod.CONFIG


# =========================
# COLUMN MAP 欄位對應
# =========================
# Maps Traditional Chinese column headers (used in CSV/Excel input files)
# to internal English keys used throughout the codebase.
# 將輸入檔案的中文欄位名稱對應為程式內部使用的英文鍵值。

COLUMN_MAP = {
    "類型": "type",
    "資料": "data",
    "WiFi名稱": "ssid",
    "WiFi密碼": "password",
    "加密方式": "security",
    "Qr-code檔案名稱": "name"
}

# Columns that must be present in every input file.
# 每個輸入檔案必須包含的欄位。
REQUIRED_COLUMNS = ["類型"]


def validate_columns(headers):
    """Raises ValueError if any required column is missing from the header row.
    若標題列缺少必要欄位則拋出例外。"""
    missing = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing:
        raise ValueError(f"❌ 缺少必要欄位: {missing}")


def normalize_row(row):
    """Translates column names via COLUMN_MAP and strips whitespace from string values.
    Non-string values (e.g. numbers from Excel) are converted to empty string if falsy.
    將欄位名稱透過 COLUMN_MAP 轉換，並去除字串值的前後空白；
    非字串型別（如 Excel 數值）若為空值則轉為空字串。"""
    return {
        COLUMN_MAP.get(k, k): v.strip() if isinstance(v, str) else (v or "")
        for k, v in row.items()
    }


# =========================
# FORMATTER QR 資料格式化
# =========================
# Converts a normalized row dict into the final string to be encoded in the QR code.
# Each type follows a specific encoding format defined by QR code standards.
# 將正規化後的資料列轉換為 QR code 編碼所需的字串。
# 各類型遵循對應的 QR code 標準編碼格式。

def format_qr_data(row):
    """Returns the QR code payload string based on the row type.
    回傳依資料類型產生的 QR code 內容字串。

    Supported types / 支援類型:
      url, text, line — uses the 'data' field directly / 直接使用 data 欄位
      wifi           — encodes as WIFI:T:...;S:...;P:;; format / 編碼為 WiFi 連線字串
    """
    t = row.get("type", "").lower()

    if t in ["url", "text", "line"]:
        return row.get("data", "")

    if t == "wifi":
        ssid = row.get("ssid", "")
        password = row.get("password", "")
        security = row.get("security", "WPA")  # Default to WPA if not specified / 未指定時預設 WPA
        return f"WIFI:T:{security};S:{ssid};P:{password};;"

    raise ValueError(f"❌ 不支援類型: {t}")


# =========================
# FILENAME STRATEGY 檔名策略
# =========================
# Priority: explicit name column > SSID (for WiFi) > auto-generated fallback.
# 優先順序：明確指定的名稱 > WiFi SSID > 自動產生的備用檔名。

def generate_filename(row, index):
    """Returns a safe filename for the QR code image.
    If no name is provided, falls back to '{type}_{index:05d}'.
    回傳 QR code 圖片的合法檔名；未指定名稱時自動產生 '{類型}_{序號}' 格式。"""
    name = row.get("name") or row.get("ssid")
    if name:
        return sanitize_filename(name)
    row_type = row.get("type", "data").lower()
    return f"{row_type}_{index:05d}"


def sanitize_filename(name):
    """Removes characters that are illegal in Windows/Unix filenames.
    移除 Windows / Unix 檔名中不合法的字元。"""
    return "".join(c for c in name if c not in r'\/:*?"<>|')


# =========================
# QR SAVE QR 圖片輸出
# =========================
# Renders the QR code and saves it to the output directory.
# For PNG/JPG, segno renders to an in-memory buffer first to avoid temporary files.
# 產生 QR code 並儲存至輸出資料夾。
# PNG/JPG 格式先渲染至記憶體緩衝區，避免產生暫存檔。

def save_qr(data, filename, config):
    """Encodes `data` as a QR code and saves it in the configured format and scale.
    將 `data` 編碼為 QR code 並依設定的格式與縮放比例儲存。"""
    qr_conf = config["qr"]
    out_conf = config["output"]
    output_dir = config["paths"]["output_dir"]

    qr = segno.make(data, error=qr_conf["error_correction"])
    fmt = out_conf["format"]
    scale = qr_conf["scale"]
    base_path = os.path.join(output_dir, filename)

    if fmt == "SVG":
        qr.save(base_path + ".svg", scale=scale)
    else:
        # Render to memory buffer to avoid writing a temporary file on disk.
        # 先輸出至記憶體，避免產生暫存檔。
        buf = io.BytesIO()
        qr.save(buf, kind="png", scale=scale)
        buf.seek(0)
        img = Image.open(buf)

        if fmt == "PNG":
            img.save(base_path + ".png", "PNG")
        elif fmt == "JPG":
            img.convert("RGB").save(base_path + ".jpg", "JPEG")


# =========================
# TXT PROCESS 純文字檔處理
# =========================
# Reads a .txt file and auto-detects whether the content is one or more URLs,
# or plain text. Each URL becomes a separate QR code; plain text becomes one.
# 讀取 .txt 檔並自動判斷內容為網址或純文字。
# 網址模式下每筆 URL 各產生一張 QR code；純文字模式則整個檔案產生一張。

# Matches a single URL token: must start with http(s):// and contain no whitespace.
# 比對單一 URL token：須以 http(s):// 開頭且不含空白字元。
_URL_TOKEN_RE = re.compile(r'^https?://[^\s]+$')

# Characters that indicate the token is not a pure URL (CJK and fullwidth ranges).
# 判定 token 不是純網址的字元集（CJK 及全形字元範圍）。
_NON_URL_CHARS_RE = re.compile(r'[\u4e00-\u9fff\u3040-\u30ff\u3400-\u4dbf\uff00-\uffef]')


def _is_url(token: str) -> bool:
    """Returns True if the token (after stripping whitespace) is a valid URL.
    Leading/trailing whitespace is tolerated as a fault-tolerance measure.
    去頭尾空白後判斷是否為合法 URL；允許前後空白以提高容錯性。"""
    token = token.strip()
    return bool(_URL_TOKEN_RE.match(token)) and not _NON_URL_CHARS_RE.search(token)


def _split_tokens(content: str) -> list[str]:
    """Splits content by newlines and commas, returning non-empty stripped tokens.
    Supports single-line comma-separated URLs and multi-line URL lists.
    以換行與逗號分割內容，回傳非空的 token 清單。
    支援單行逗號分隔及多行換行分隔兩種格式。"""
    tokens = []
    for line in content.splitlines():
        for part in line.split(","):
            t = part.strip()
            if t:
                tokens.append(t)
    return tokens


def process_txt(file_path, config):
    """Processes a .txt file: detects URLs vs plain text and generates QR code(s).
    處理 .txt 檔：自動偵測網址或純文字並產生對應 QR code。

    - All tokens are URLs → one QR per URL, named '{stem}' or '{stem}_{index:05d}'
    - Any non-URL token found → entire file treated as plain text, one QR total
    - 全部為網址 → 每筆 URL 各一張，檔名為 '{stem}' 或 '{stem}_{序號}'
    - 含非網址內容 → 整個檔案當文本，產生一張 QR code
    """
    with open(file_path, encoding="utf-8") as f:
        content = f.read().strip()

    if not content:
        return

    basename = os.path.basename(file_path)
    stem = sanitize_filename(os.path.splitext(basename)[0])
    tokens = _split_tokens(content)

    url_flags = [_is_url(t) for t in tokens]

    if all(url_flags):
        logging.info(f"  判斷為 URL（{len(tokens)} 筆）: {basename}")
        if len(tokens) == 1:
            save_qr(tokens[0], stem, config)
        else:
            for i, url in enumerate(tokens, 1):
                save_qr(url, f"{stem}_{i:05d}", config)
    else:
        # Report the first token that caused the file to be treated as plain text.
        # 回報導致判定為文本的第一個非網址 token。
        first_non_url = next(t for t, ok in zip(tokens, url_flags) if not ok)
        logging.info(f"  判斷為文本（含非網址內容：「{first_non_url[:40]}」）: {basename}")
        save_qr(content, stem, config)


# =========================
# CSV PROCESS CSV 檔處理
# =========================
# Reads a CSV file using the delimiter from config and generates one QR per data row.
# 依設定的分隔符號讀取 CSV 檔，每筆資料列產生一張 QR code。

def process_csv(file_path, config):
    """Iterates over all data rows in a CSV file and generates a QR code per row.
    逐列讀取 CSV 資料並為每列產生 QR code；單列錯誤不中斷整體流程。"""
    delimiter = config.get("csv", {}).get("delimiter", ",")
    with open(file_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=delimiter)

        validate_columns(reader.fieldnames)

        for i, row in enumerate(reader):
            try:
                row = normalize_row(row)
                data = format_qr_data(row)

                if not data:
                    continue

                filename = generate_filename(row, i + 1)
                save_qr(data, filename, config)

            except Exception as e:
                logging.error(f"CSV 第{i+1}筆: {e}")


# =========================
# EXCEL PROCESS Excel 檔處理
# =========================
# Reads an .xlsx file and generates one QR per data row (row 1 is treated as header).
# Cell values are read individually to handle openpyxl format-code/type mismatches.
# 讀取 .xlsx 檔，第一列視為標題，其餘每列產生一張 QR code。
# 逐格讀取以應對 openpyxl 格式代碼與儲存格型別不符的已知問題。

def _read_excel_rows(sheet):
    """Reads all cell values row by row, catching per-cell errors gracefully.
    openpyxl may raise errors when a date-formatted cell contains a string value;
    this wrapper converts such cells to None instead of crashing.
    逐格讀取所有儲存格值，個別捕捉例外。
    當日期格式儲存格內含字串值時 openpyxl 會報錯，此函式將問題格轉為 None 而非中止。"""
    result = []
    for row in sheet.iter_rows():
        values = []
        for cell in row:
            try:
                values.append(cell.value)
            except Exception:
                values.append(None)
        result.append(tuple(values))
    return result


def process_excel(file_path, config):
    """Iterates over all data rows in an Excel file and generates a QR code per row.
    逐列讀取 Excel 資料並為每列產生 QR code；單列錯誤不中斷整體流程。"""
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    if sheet is None:
        raise ValueError("❌ Excel 無法讀取工作表")
    rows = _read_excel_rows(sheet)
    if not rows:
        return
    headers = rows[0]

    validate_columns(headers)

    for i, row_data in enumerate(rows[1:]):
        try:
            row = dict(zip(headers, row_data))
            row = normalize_row(row)

            data = format_qr_data(row)
            if not data:
                continue

            filename = generate_filename(row, i + 1)
            save_qr(data, filename, config)

        except Exception as e:
            logging.error(f"Excel 第{i+1}筆: {e}")


# =========================
# LOGGING 執行紀錄
# =========================
# Clears the log directory on each run to keep only the latest log file.
# Both file and console handlers are attached so output is visible in the terminal.
# 每次執行前清空 log 資料夾，僅保留本次紀錄。
# 同時輸出至檔案與終端機，方便即時監控與事後查閱。

def setup_logging(log_dir: str) -> str:
    """Clears the log directory, creates a timestamped log file, and configures logging.
    Returns the path of the newly created log file.
    清空 log 資料夾，建立帶時間戳的 log 檔並設定 logging，回傳 log 檔路徑。"""
    if os.path.exists(log_dir):
        shutil.rmtree(log_dir)
    os.makedirs(log_dir)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join(log_dir, f"run_{timestamp}.log")

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )
    return log_path


# =========================
# MAIN 主程式
# =========================
# Entry point. Loads config, sets up logging, then walks the input directory
# recursively and dispatches each supported file to the appropriate handler.
# 程式進入點。載入設定、初始化 log，遞迴掃描輸入資料夾，
# 並將各支援格式的檔案分派至對應的處理函式。

def main():
    config = load_config()
    input_dir = config["paths"]["input_dir"]
    output_dir = config["paths"]["output_dir"]
    log_dir = config["paths"].get("log_dir", "logs")

    os.makedirs(output_dir, exist_ok=True)
    log_path = setup_logging(log_dir)
    logging.info(f"Log 檔案 / Log file: {log_path}")

    # Map file extensions to their processor functions.
    # 副檔名與對應處理函式的對照表。
    handlers = {
        ".csv": process_csv,
        ".xlsx": process_excel,
        ".txt": process_txt,
    }

    for dirpath, _, filenames in os.walk(input_dir):
        for file in filenames:
            ext = os.path.splitext(file)[1].lower()
            handler = handlers.get(ext)
            if handler is None:
                continue  # Skip unsupported file types. / 略過不支援的檔案格式。
            file_path = os.path.join(dirpath, file)
            logging.info(f"處理 / Processing: {os.path.relpath(file_path, input_dir)}")
            try:
                handler(file_path, config)
            except Exception as e:
                logging.error(f"{file}: {e}")

    logging.info("✅ 全部完成 / All done!")


if __name__ == "__main__":
    main()
